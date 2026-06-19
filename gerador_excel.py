"""
gerador_excel.py
----------------
Gera o Excel de distribuição de terraplenagem com as abas:
  3_RESUMO_CA  — resumo de cortes, aterros e CL por ramo
  4_QDM        — quadro de distribuição de material
  5_RESUMO_QDM — resumo com faixas de DMT, compactação e momento

Segue exatamente o padrão do arquivo de referência CMAT-SEGMENTO_B_rev.xlsx.
Usa fórmulas SUMIFS no 5_RESUMO_QDM referenciando o 4_QDM.
Duas colunas de flag:
  col U = flag_nao_soma_corte  (1 = linha vem de outro ramo — não soma corte neste ramo)
  col V = flag_nao_soma_aterro (1 = linha vai para outro ramo — não soma aterro neste ramo)
"""

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               numbers)
from openpyxl.utils import get_column_letter
from typing import List, Dict
from distribuidor2 import LinhaQDM, ResultadoDistribuicao
from detector_trechos import Trecho, TipoTrecho, ResultadoDeteccao

# Cores padrão
COR_CABECALHO  = "00366092"   # azul escuro
COR_SUBHEADER  = "00DCE6F1"   # azul claro
COR_CORTE      = "00FFC7CE"   # rosa claro
COR_ATERRO     = "00C6EFCE"   # verde claro
COR_CL         = "00FFEB9C"   # amarelo claro
COR_BF         = "00E2EFDA"   # verde pálido
COR_AE         = "00E8D5B7"   # marrom claro
FONTE_BRANCA   = "FFFFFFFF"
FONTE_PADRAO   = "FF000000"

LINHA_INICIO_DADOS = 14  # linha onde começam os dados (igual ao modelo)

# Faixas de DMT (em km)
FAIXAS_DMT = [
    (0,      0.05,  "0 < DMT ≤ 50 m"),
    (0.05,   0.2,   "50 < DMT ≤ 200 m"),
    (0.2,    0.4,   "200 < DMT ≤ 400 m"),
    (0.4,    0.6,   "400 < DMT ≤ 600 m"),
    (0.6,    0.8,   "600 < DMT ≤ 800 m"),
    (0.8,    1.0,   "800 < DMT ≤ 1000 m"),
    (1.0,    1.2,   "1000 < DMT ≤ 1200 m"),
    (1.2,    1.4,   "1200 < DMT ≤ 1400 m"),
    (1.4,    1.6,   "1400 < DMT ≤ 1600 m"),
    (1.6,    1.8,   "1600 < DMT ≤ 1800 m"),
    (1.8,    2.0,   "1800 < DMT ≤ 2000 m"),
    (2.0,    2.5,   "2000 < DMT ≤ 2500 m"),
    (2.5,    3.0,   "2500 < DMT ≤ 3000 m"),
    (3.0,    None,  "DMT > 3000 m"),
]


# ---------------------------------------------------------------------------
# Helpers de estilo
# ---------------------------------------------------------------------------

def estilo_cabecalho(cell, texto=None, negrito=True, cor_fundo=COR_CABECALHO,
                     cor_fonte=FONTE_BRANCA, centralizar=True):
    if texto is not None:
        cell.value = texto
    cell.font = Font(name="Arial", bold=negrito, color=cor_fonte, size=9)
    cell.fill = PatternFill("solid", fgColor=cor_fundo)
    if centralizar:
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def estilo_dado(cell, valor=None, formato=None, negrito=False,
                cor_fundo=None, centralizar=False):
    if valor is not None:
        cell.value = valor
    cell.font = Font(name="Arial", bold=negrito, size=9)
    if cor_fundo:
        cell.fill = PatternFill("solid", fgColor=cor_fundo)
    if formato:
        cell.number_format = formato
    if centralizar:
        cell.alignment = Alignment(horizontal="center", vertical="center")


def borda_fina():
    lado = Side(style="thin")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def aplicar_borda(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = borda_fina()


# ---------------------------------------------------------------------------
# Aba 3_RESUMO_CA
# ---------------------------------------------------------------------------

def gerar_resumo_ca(wb: Workbook,
                    resultados: List[ResultadoDeteccao],
                    nome_projeto: str = "",
                    fatores_hom: dict = None,
                    restricoes=None,
                    mapeamento=None):
    if fatores_hom is None:
        fatores_hom = {}

    def fh_cat(categoria: str) -> float:
        """
        Retorna FH da categoria.
        Usa o mapeamento para identificar quais materiais pertencem a cada
        categoria — evita depender do nome do material (ex: 'corpo' vs 'aterro').
        Fallback: busca por palavras-chave no nome do material.
        """
        mats_c1 = list(getattr(mapeamento, 'corte1',    [])) if mapeamento else []
        mats_c2 = list(getattr(mapeamento, 'corte2',    [])) if mapeamento else []
        mats_c3 = list(getattr(mapeamento, 'corte3',    [])) if mapeamento else []
        mats_ca = list(getattr(mapeamento, 'aterro_ca', [])) if mapeamento else []
        mats_cf = list(getattr(mapeamento, 'aterro_cf', [])) if mapeamento else []
        mapa_cat = {'C1': mats_c1, 'C2': mats_c2, 'C3': mats_c3,
                    'CA': mats_ca, 'CF': mats_cf}
        for mat in mapa_cat.get(categoria, []):
            if mat in fatores_hom:
                return fatores_hom[mat]
        for mat, fh in fatores_hom.items():
            m = mat.upper().replace(' ', '').replace('ª', 'A').replace('°', 'A')
            if categoria == 'C1' and 'CORTE1' in m: return fh
            if categoria == 'C2' and 'CORTE2' in m: return fh
            if categoria == 'C3' and 'CORTE3' in m: return fh
            if categoria == 'CA' and 'ATERRO' in m and 'CF' not in m: return fh
            if categoria == 'CF' and ('CF' in m or 'FINAL' in m): return fh
        return 1.0

    ws = wb.create_sheet("3_RESUMO_CA")

    # Cabeçalho geral (linhas 1-8)
    ws.merge_cells("A6:F6")
    ws["A6"] = nome_projeto or "MEMORIAL DE CÁLCULO DE VOLUMES"
    ws["A6"].font = Font(name="Arial", bold=True, size=10)

    ws.merge_cells("A7:F7")
    ws["A7"] = "RESUMO CORTE E ATERRO"
    ws["A7"].font = Font(name="Arial", bold=True, size=10)

    # Cabeçalho da tabela (linhas 9-10)
    headers_L9 = [
        ("A9", "TIPO"), ("B9", "Nº"), ("C9", "KM INICIO"),
        ("D9", "CMv"),  ("E9", "KM FINAL"), ("F9", "EXTENSÃO m"),
    ]
    for addr, txt in headers_L9:
        estilo_cabecalho(ws[addr], txt)

    # Grupos de volume
    ws.merge_cells("G9:M9")
    estilo_cabecalho(ws["G9"], "VOLUME GEOMÉTRICO m³")
    ws.merge_cells("N9:T9")
    estilo_cabecalho(ws["N9"], "VOLUME HOMOGENEIZADO m³")
    ws.merge_cells("U9:U10")
    estilo_cabecalho(ws["U9"], "C.L")

    # Sub-cabeçalhos linha 10
    cols_vol = ["Corte 1ª","Corte 2ª","Corte 3ª","Total Cortes",
                "Aterro CA","Aterro CF","Total Aterros"]
    for j, txt in enumerate(cols_vol):
        # Geométrico: cols G-M (7-13)
        estilo_cabecalho(ws.cell(10, 7+j), txt)
        # Homogeneizado: cols N-T (14-20)
        estilo_cabecalho(ws.cell(10, 14+j), txt)

    # Larguras
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 10
    for col in range(7, 22):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # Dados por ramo
    row = LINHA_INICIO_DADOS

    for res in resultados:
        # Título do ramo
        ws.merge_cells(f"A{row}:U{row}")
        ws[f"A{row}"] = res.ramo
        ws[f"A{row}"].font = Font(name="Arial", bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=COR_SUBHEADER)
        row += 1

        # Separar trechos por tipo
        cortes = [t for t in res.trechos if t.tipo == TipoTrecho.CORTE]
        aterros_ca = [t for t in res.trechos
                      if t.tipo == TipoTrecho.ATERRO and t.categoria == "CA"]
        aterros_cf = [t for t in res.trechos
                      if t.tipo == TipoTrecho.ATERRO and t.categoria == "CF"]
        cls = [t for t in res.trechos if t.tipo == TipoTrecho.CL]

        def escrever_trecho(t: Trecho, tipo_label: str, cor: str,
                            label_override: str = None):
            nonlocal row
            ws[f"A{row}"] = tipo_label
            ws[f"B{row}"] = label_override if label_override else t.label
            ws[f"C{row}"] = t.estaca_ini
            ws[f"D{row}"] = t.cmv_label
            ws[f"E{row}"] = t.estaca_fin
            # Extensão como fórmula (igual ao modelo)
            ws[f"F{row}"] = t.extensao

            # Vol geométrico e homogeneizado
            vg = t.volumes_geo if hasattr(t, 'volumes_geo') and t.volumes_geo else {}
            vol_hom = t.vol_total_hom

            # vol_geo: se volumes_geo preenchido e diferente do hom → usa direto
            # senão calcula vol_geo = vol_hom × FH
            # CL é material de C1 compensado lateralmente → usa FH_C1
            if vg and abs(vg.get("total", vol_hom) - vol_hom) > 0.01:
                vol_geo = vg.get("total", vol_hom)
            else:
                cat = t.categoria
                fh_geo = fh_cat('C1') if cat == 'CL' else fh_cat(cat)
                vol_geo = round(vol_hom * fh_geo, 4)

            if t.tipo == TipoTrecho.CORTE:
                cat = t.categoria
                c1g = vol_geo if cat == "C1" else 0
                c2g = vol_geo if cat == "C2" else 0
                c3g = vol_geo if cat == "C3" else 0
                c1h = vol_hom if cat == "C1" else 0
                c2h = vol_hom if cat == "C2" else 0
                c3h = vol_hom if cat == "C3" else 0
                ws[f"G{row}"] = c1g; ws[f"H{row}"] = c2g; ws[f"I{row}"] = c3g
                ws[f"J{row}"] = f"=SUM(G{row}:I{row})"
                ws[f"K{row}"] = 0; ws[f"L{row}"] = 0
                ws[f"M{row}"] = f"=SUM(K{row}:L{row})"
                ws[f"N{row}"] = c1h; ws[f"O{row}"] = c2h; ws[f"P{row}"] = c3h
                ws[f"Q{row}"] = f"=SUM(N{row}:P{row})"
                ws[f"R{row}"] = 0; ws[f"S{row}"] = 0
                ws[f"T{row}"] = f"=S{row}+R{row}"
            elif t.tipo == TipoTrecho.ATERRO:
                ca_geo = vol_geo if t.categoria == "CA" else 0
                cf_geo = vol_geo if t.categoria == "CF" else 0
                ca_hom = vol_hom if t.categoria == "CA" else 0
                cf_hom = vol_hom if t.categoria == "CF" else 0
                ws[f"G{row}"] = 0; ws[f"H{row}"] = 0; ws[f"I{row}"] = 0
                ws[f"J{row}"] = f"=SUM(G{row}:I{row})"
                ws[f"K{row}"] = ca_geo; ws[f"L{row}"] = cf_geo
                ws[f"M{row}"] = f"=SUM(K{row}:L{row})"
                ws[f"N{row}"] = 0; ws[f"O{row}"] = 0; ws[f"P{row}"] = 0
                ws[f"Q{row}"] = f"=SUM(N{row}:P{row})"
                ws[f"R{row}"] = ca_hom; ws[f"S{row}"] = cf_hom
                ws[f"T{row}"] = f"=S{row}+R{row}"
            elif t.tipo == TipoTrecho.CL:
                ws[f"G{row}"] = vol_geo; ws[f"H{row}"] = 0; ws[f"I{row}"] = 0
                ws[f"J{row}"] = f"=SUM(G{row}:I{row})"
                ws[f"K{row}"] = 0; ws[f"L{row}"] = 0
                ws[f"M{row}"] = f"=SUM(K{row}:L{row})"
                ws[f"N{row}"] = vol_hom; ws[f"O{row}"] = 0; ws[f"P{row}"] = 0
                ws[f"Q{row}"] = f"=SUM(N{row}:P{row})"
                ws[f"R{row}"] = 0; ws[f"S{row}"] = 0
                ws[f"T{row}"] = f"=S{row}+R{row}"
                ws[f"U{row}"] = vol_hom

            # Formatar números
            for col in range(7, 22):
                ws.cell(row, col).number_format = "#,##0.00"

            # Cor por tipo
            for col in range(1, 22):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=cor)
                ws.cell(row, col).font = Font(name="Arial", size=9)
                ws.cell(row, col).border = borda_fina()

            row += 1

        # Cortes
        for t in cortes:
            escrever_trecho(t, "Corte", COR_CORTE)
        # CF primeiro (como no modelo)
        for t in aterros_cf:
            escrever_trecho(t, "Aterro", COR_ATERRO)
        # CA
        for t in aterros_ca:
            escrever_trecho(t, "Aterro", COR_ATERRO)
        # CL — se ramo tem corte_somente_bf, escreve como Corte (vai para BF)
        ramo_somente_bf = (restricoes is not None and
                           restricoes.tem_restricoes() and
                           restricoes.corte_somente_bf(res.ramo))
        for t in cls:
            if ramo_somente_bf:
                escrever_trecho(t, "Corte", COR_CORTE,
                                label_override=t.label.replace('CL-', 'C1-'))
            else:
                escrever_trecho(t, "C. Lateral", COR_CL)

        row += 1  # linha em branco entre ramos

    return ws


# ---------------------------------------------------------------------------
# Aba 4_QDM
# ---------------------------------------------------------------------------

def gerar_qdm(wb: Workbook,
              resultado: ResultadoDistribuicao,
              resultados_deteccao: List[ResultadoDeteccao]) -> int:
    """Gera a aba QDM. Retorna a última linha de dados."""
    ws = wb.create_sheet("4_QDM")

    # Cabeçalho linha 9
    cab9 = [
        "ORIGEM MATERIAL ESCAVADO [CORTE] m³", None, None, None,
        None, None, None, None,
        "DMT km", None, None,
        "MMT  m³xkm", None, None,
        "DESTINO MATERIAL ESCAVADO [ATERRO] m³", None, None, None, None,
        "OBS.", None, None
    ]
    for j, txt in enumerate(cab9):
        if txt:
            ws.cell(9, j+1).value = txt
            ws.cell(9, j+1).font = Font(name="Arial", bold=True, size=9)
            ws.cell(9, j+1).fill = PatternFill("solid", fgColor=COR_CABECALHO)
            ws.cell(9, j+1).font = Font(name="Arial", bold=True,
                                         color=FONTE_BRANCA, size=9)
            ws.cell(9, j+1).alignment = Alignment(horizontal="center",
                                                    vertical="center",
                                                    wrap_text=True)

    # Sub-cabeçalho linha 10
    # A   B          C      D        E         F         G         H
    # Corte KmIni  CMv   KmFinal  Corte1ª  Corte2ª  Corte3ª  Total
    # I     J      K     L        M         N         O         P
    # DT  Fixa   Total  MMT C1  MMT C2   MMT C3   Aterro   KmIni
    # Q      R        S      T     U                  V
    # CMv   KmFinal  Local  OBS  flag_corte         flag_aterro
    sub10 = [
        "Corte","Km Inicio","CMv","Km Final",
        "Corte 1ª","Corte 2ª","Corte 3ª","Total",
        "DT","Fixa","Total",
        "Corte 1ª","Corte 2ª","Corte 3ª",
        "Aterro","Km Inicio","CMv","Km Final","Local",
        "OBS.",None,
        "Flag\nCorte","Flag\nAterro"
    ]
    for j, txt in enumerate(sub10):
        c = ws.cell(10, j+1)
        estilo_cabecalho(c, txt)

    # Larguras de coluna
    larguras = {
        1:10, 2:13, 3:13, 4:13,
        5:12, 6:12, 7:12, 8:12,
        9:8, 10:7, 11:8,
        12:12, 13:12, 14:12,
        15:10, 16:13, 17:13, 18:13, 19:6,
        20:25, 21:1,
        22:6, 23:6
    }
    for col, w in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Agrupar linhas por ramo
    # Ordem: usa resultados_deteccao para manter a mesma ordem do RESUMO_CA
    # Fallback: ordem de aparição nas linhas para ramos não detectados
    ramos_deteccao = [r.ramo for r in resultados_deteccao] if resultados_deteccao else []
    ramos_vistos = []
    for l in resultado.linhas_qdm:
        if l.flag_nao_soma_corte == 1:
            ramo_ref = l.ramo_destino
        else:
            ramo_ref = l.ramo_origem if l.ramo_origem else l.ramo_destino
        if ramo_ref and ramo_ref not in ramos_vistos:
            ramos_vistos.append(ramo_ref)

    # Ordenar pela ordem da detecção, mantendo ramos extras no final
    ramos_ordem = [r for r in ramos_deteccao if r in ramos_vistos]
    for r in ramos_vistos:
        if r not in ramos_ordem:
            ramos_ordem.append(r)

    row = LINHA_INICIO_DADOS

    for ramo in ramos_ordem:
        # Título do ramo
        ws.merge_cells(f"A{row}:T{row}")
        ws[f"A{row}"] = ramo
        ws[f"A{row}"].font = Font(name="Arial", bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=COR_SUBHEADER)
        row += 1

        # Linhas do ramo:
        # - origem no ramo (cortes, CL, BF, vai_para) — excluir vem_do (flag_nao_soma_corte=1)
        # - destino no ramo com origem vazia (empréstimos sem cadastro)
        # - destino no ramo vindo de outro ramo (flag_nao_soma_corte=1)
        linhas_ramo = [
            l for l in resultado.linhas_qdm
            if (l.ramo_origem == ramo and l.flag_nao_soma_corte != 1)  # cortes/CL/BF/vai_para do ramo
            or (not l.ramo_origem and l.ramo_destino == ramo)           # AE sem cadastro
            or (l.ramo_destino == ramo and l.flag_nao_soma_corte == 1)  # vem_do — material externo chegando
        ]

        for l in linhas_ramo:
            # Cor por tipo
            if l.tipo_destino == "CL":
                cor = COR_CL
            elif l.tipo_destino == "BF":
                cor = COR_BF
            elif l.tipo_destino in ("CA", "CF"):
                cor = COR_ATERRO if not l.flag_nao_soma_corte else "00FFFF99"
            else:
                cor = "00FFFFFF"

            fill = PatternFill("solid", fgColor=cor)
            font = Font(name="Arial", size=9)

            def sc(col, val, fmt=None):
                c = ws.cell(row, col)
                c.value = val
                c.font = font
                c.fill = fill
                c.border = borda_fina()
                c.alignment = Alignment(horizontal="center", vertical="center")
                if fmt:
                    c.number_format = fmt

            # Determinar se é empréstimo (AE): volume vai nas colunas de aterro
            is_ae = not l.ramo_origem and l.tipo_destino in ("CA", "CF")

            # --- ORIGEM ---
            sc(1, l.label_origem)
            sc(2, l.estaca_ini_origem)
            sc(3, l.cmv_origem)
            sc(4, l.estaca_fin_origem)

            if is_ae:
                # AE: colunas de corte ficam vazias, volume aparece no destino
                sc(5, None, "#,##0.00")
                sc(6, None, "#,##0.00")
                sc(7, None, "#,##0.00")
            else:
                sc(5, l.vol_c1, "#,##0.00")
                sc(6, l.vol_c2, "#,##0.00")
                sc(7, l.vol_c3, "#,##0.00")

            ws.cell(row, 8).value  = f"=SUM(E{row}:G{row})"
            ws.cell(row, 8).font   = font
            ws.cell(row, 8).fill   = fill
            ws.cell(row, 8).border = borda_fina()
            ws.cell(row, 8).number_format = "#,##0.00"

            # --- DMT ---
            dmt_var = l.dmt_var if l.dmt_total > 0 else None
            dmt_fixa = l.dmt_fixa if l.dmt_total > 0 else None
            sc(9,  dmt_var,  "0.000")
            sc(10, dmt_fixa, "0.000")
            if l.dmt_total > 0:
                ws.cell(row, 11).value = f"=I{row}+J{row}"
            else:
                ws.cell(row, 11).value = None
            ws.cell(row, 11).font   = font
            ws.cell(row, 11).fill   = fill
            ws.cell(row, 11).border = borda_fina()
            ws.cell(row, 11).number_format = "0.000"

            # --- MMT ---
            ws.cell(row, 12).value  = f"=K{row}*E{row}"
            ws.cell(row, 13).value  = f"=K{row}*F{row}"
            ws.cell(row, 14).value  = f"=K{row}*G{row}"
            for col in [12, 13, 14]:
                ws.cell(row, col).font   = font
                ws.cell(row, col).fill   = fill
                ws.cell(row, col).border = borda_fina()
                ws.cell(row, col).number_format = "#,##0.00"

            # --- DESTINO ---
            sc(15, l.label_destino)
            sc(16, l.estaca_ini_destino)
            sc(17, l.cmv_destino)
            sc(18, l.estaca_fin_destino)
            sc(19, l.tipo_destino)

            # Volume de aterro no destino (para AE)
            if is_ae:
                # AE: volume no total, OBS com nome do empréstimo
                ws.cell(row, 8).value  = round(l.vol_total, 4)
                ws.cell(row, 8).number_format = "#,##0.00"
                sc(20, l.obs)  # "vem do empréstimo" ou nome da jazida
            else:
                sc(20, l.obs)

            # --- FLAGS ---
            sc(22, l.flag_nao_soma_corte  if hasattr(l, 'flag_nao_soma_corte')  else 0)
            sc(23, l.flag_nao_soma_aterro if hasattr(l, 'flag_nao_soma_aterro') else 0)

            row += 1

        row += 1  # separador entre ramos

    return row - 1  # última linha de dados


# ---------------------------------------------------------------------------
# Aba 5_RESUMO_QDM
# ---------------------------------------------------------------------------

def gerar_resumo_qdm(wb: Workbook, ultima_linha_qdm: int,
                     fatores_hom: dict = None,
                     mapeamento=None):
    if fatores_hom is None:
        fatores_hom = {}

    # FHs por categoria
    def _fh(categoria: str) -> float:
        mats_c1 = list(getattr(mapeamento, 'corte1',    [])) if mapeamento else []
        mats_c2 = list(getattr(mapeamento, 'corte2',    [])) if mapeamento else []
        mats_c3 = list(getattr(mapeamento, 'corte3',    [])) if mapeamento else []
        mats_ca = list(getattr(mapeamento, 'aterro_ca', [])) if mapeamento else []
        mats_cf = list(getattr(mapeamento, 'aterro_cf', [])) if mapeamento else []
        mapa_cat = {'C1': mats_c1, 'C2': mats_c2, 'C3': mats_c3,
                    'CA': mats_ca, 'CF': mats_cf}
        for mat in mapa_cat.get(categoria, []):
            if mat in fatores_hom:
                return fatores_hom[mat]
        for mat, fh in fatores_hom.items():
            m = mat.upper().replace(' ', '').replace('ª', 'A').replace('°', 'A')
            if categoria == 'C1' and 'CORTE1' in m: return fh
            if categoria == 'C2' and 'CORTE2' in m: return fh
            if categoria == 'C3' and 'CORTE3' in m: return fh
            if categoria == 'CA' and 'ATERRO' in m and 'CF' not in m: return fh
            if categoria == 'CF' and ('CF' in m or 'FINAL' in m): return fh
        return 1.0

    fh_c1 = _fh('C1')
    fh_c2 = _fh('C2')
    fh_c3 = _fh('C3')
    fh_ca = _fh('CA')
    fh_cf = _fh('CF')

    ws = wb.create_sheet("5_RESUMO_QDM")
    ini = LINHA_INICIO_DADOS
    fim = ultima_linha_qdm
    q   = "'4_QDM'"

    def r(col):
        """Helper: range reference no QDM."""
        return f"{q}!${col}${ini}:${col}${fim}"

    H = r("H"); F = r("F"); G = r("G")
    K = r("K"); L = r("L"); M = r("M"); N = r("N")
    S = r("S"); A = r("A"); W = r("V"); V = r("W")

    # Título
    ws.merge_cells("A8:G8")
    ws["A8"] = "RESUMO DO QUADRO DE DISTRIBUIÇÃO DE MATERIAL"
    ws["A8"].font = Font(name="Arial", bold=True, size=11)

    # Cabeçalho origem/destino
    ws.merge_cells("A10:C10")
    estilo_cabecalho(ws["A10"], "ORIGEM GEOMÉTRICO")
    ws.merge_cells("D10:H10")
    estilo_cabecalho(ws["D10"], "DESTINO GEOMÉTRICO")
    for j, txt in enumerate(["CORTES INTERPERFIS","EMPRÉSTIMO","TOTAL",
                              "CORPO DE ATERRO","CAMADA FINAL","ADME","TOTAL"]):
        estilo_cabecalho(ws.cell(11, j+1), txt)

    # Fórmulas linha 14 — volumes geométricos = vol_hom × FH
    # E=vol_c1, F=vol_c2, G=vol_c3 no QDM
    E = r("E"); F2 = r("F"); G2 = r("G")

    # A14 — Cortes interperfis (C1+C2+C3+CL) geométrico
    ws["A14"] = (
        f'=SUMIFS({H},{A},"=C-*",{W},"<>1")*{fh_c1}'
        f'+SUMIFS({H},{A},"=C2-*",{W},"<>1")*{fh_c2}'
        f'+SUMIFS({H},{A},"=C3-*",{W},"<>1")*{fh_c3}'
        f'+SUMIFS({H},{A},"=CL-*",{W},"<>1")*{fh_c1}'
    )
    # B14 — Empréstimo (AE) — usa FH_C1 como fallback
    ws["B14"] = f'=SUMIFS({H},{A},"=AE-*")*{fh_c1}'
    ws["C14"] = "=A14+B14"

    # D14 — Destino CA — CL destino usa FH_CA (não FH_C1)
    ws["D14"] = (
        f'=SUMIFS({H},{S},"=CA",{V},"<>1")*{fh_ca}'
        f'+SUMIFS({H},{S},"=CL",{V},"<>1")*{fh_ca}'
    )
    # E14 — Destino CF
    ws["E14"] = f'=SUMIFS({H},{S},"=CF",{V},"<>1")*{fh_cf}'

    # F14 — ADME/BF — C1×FH_C1 + C2×FH_C2 + C3×FH_C3
    ws["F14"] = (
        f'=SUMIFS({H},{S},"=BF",{E},">0")*{fh_c1}'
        f'+SUMIFS({H},{S},"=BF",{F2},">0")*{fh_c2}'
        f'+SUMIFS({H},{S},"=BF",{G2},">0")*{fh_c3}'
    )
    ws["G14"] = "=D14+E14+F14"

    for col in range(1, 8):
        ws.cell(14, col).number_format = "#,##0.00"
        ws.cell(14, col).font = Font(name="Arial", bold=True, size=9)
        ws.cell(14, col).border = borda_fina()

    ws["A16"] = "OBS.:"
    ws["A16"].font = Font(name="Arial", bold=True, size=9)
    obs_txt = (f"Fator de Homogeneização aplicado: "
               f"1ª Cat. Fh={str(fh_c1).replace('.',',')}  "
               f"2ª Cat. Fh={str(fh_c2).replace('.',',')}  "
               f"3ª Cat. Fh={str(fh_c3).replace('.',',')}  "
               f"Aterro CA Fh={str(fh_ca).replace('.',',')}  "
               f"Aterro CF Fh={str(fh_cf).replace('.',',')}")
    ws["A17"] = obs_txt
    ws["A17"].font = Font(name="Arial", italic=True, size=9)

    # Escavação geométrico
    ws.merge_cells("A19:G19")
    estilo_cabecalho(ws["A19"], "ESCAVAÇÃO GEOMÉTRICO")
    for addr, txt in [("A20","ÍTEM"),("D20","UNIDADE"),("E20","QUANTIDADE"),
                      ("E21","1ª CAT"),("F21","2ª CAT"),("G21","3ª CAT")]:
        ws[addr] = txt
        ws[addr].font = Font(name="Arial", bold=True, size=9)

    row = 22
    for (dmin, dmax, label) in FAIXAS_DMT:
        ws.cell(row, 1).value = "Escavação, carga e transporte"
        ws.cell(row, 3).value = label
        ws.cell(row, 4).value = "m³"
        if dmax is None:
            crit = f'{K},">3,0"'
        else:
            lo = str(dmin).replace(".", ",")
            hi = str(dmax).replace(".", ",")
            crit = f'{K},">{lo}",{K},"<={hi}"'

        # Col5 = C1 geo: usa coluna E (vol_c1) × FH_C1 + CL × FH_C1
        ws.cell(row, 5).value = (
            f'=SUMIFS({E},{crit},{W},"<>1",{A},"=C-*")*{fh_c1}'
            f'+SUMIFS({H},{crit},{W},"<>1",{A},"=CL-*")*{fh_c1}'
        )
        # Col6 = C2 geo: usa coluna F (vol_c2) × FH_C2
        ws.cell(row, 6).value = (
            f'=SUMIFS({F2},{crit},{W},"<>1",{A},"=C2-*")*{fh_c2}'
        )
        # Col7 = C3 geo: usa coluna G (vol_c3) × FH_C3
        ws.cell(row, 7).value = (
            f'=SUMIFS({G2},{crit},{W},"<>1",{A},"=C3-*")*{fh_c3}'
        )

        for col in [1,3,4,5,6,7]:
            ws.cell(row,col).font   = Font(name="Arial", size=9)
            ws.cell(row,col).border = borda_fina()
        for col in [5,6,7]:
            ws.cell(row,col).number_format = "#,##0.00"
        row += 1

    # Momento de transporte
    ws.cell(row,1).value = "Momento de transporte"
    ws.cell(row,3).value = "DMT > 3000 m"
    ws.cell(row,4).value = "m³xkm"
    ws.cell(row,5).value = f'=SUMIFS({L},{K},">3,0",{W},"<>1")'
    ws.cell(row,6).value = f'=SUMIFS({M},{K},">3,0")'
    ws.cell(row,7).value = f'=SUMIFS({N},{K},">3,0")'
    for col in [1,3,4,5,6,7]:
        ws.cell(row,col).font   = Font(name="Arial", bold=True, size=9)
        ws.cell(row,col).border = borda_fina()
    for col in [5,6,7]:
        ws.cell(row,col).number_format = "#,##0.00"
    row += 2

    # Compactação
    ws.merge_cells(f"A{row}:G{row}")
    estilo_cabecalho(ws[f"A{row}"], "COMPACTAÇÃO")
    row += 1
    for addr, txt in [(f"A{row}","ITEM"),(f"D{row}","UNIDADE"),(f"E{row}","QUANTIDADE")]:
        ws[addr] = txt
        ws[addr].font = Font(name="Arial", bold=True, size=9)
    row += 1

    for desc, unid, formula in [
        ("Compactação de Corpo de Aterro a 100% do Proctor normal",     "m³", "=D14"),
        ("Compactação de Camada Final a 100% do Proctor intermediário",  "m³", "=E14"),
        ("Construção de corpo de aterro com material de 3ª categoria oriundo de corte - 4 passadas",
         "m³", f'=SUMIFS({H},{S},"=CA",{G2},">0",{W},"<>1")/{fh_c3}'),
        ("Regularização de bota-espera com espalhamento e compactação",  "m³", "=F14"),
    ]:
        ws.cell(row,1).value = desc
        ws.cell(row,4).value = unid
        ws.cell(row,5).value = formula
        ws.cell(row,5).number_format = "#,##0.00"
        for col in [1,4,5]:
            ws.cell(row,col).font   = Font(name="Arial", size=9)
            ws.cell(row,col).border = borda_fina()
        row += 1

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    for col in ["E","F","G"]:
        ws.column_dimensions[col].width = 15

    return ws


# ---------------------------------------------------------------------------
# Aba 1_CV EIXOS
# ---------------------------------------------------------------------------

def gerar_cv_eixos(wb: Workbook, projeto, mapeamento, nome_projeto: str = "",
                   fatores_hom: dict = None, restricoes=None):
    """
    Gera aba 1_CV EIXOS com dados por estaca de cada ramo.
    Colunas: A=Estaca, B-F=Área, G-K=Vol Geo, L-P=Vol Hom, Q=CL, R-V=Acumulado, W=OM, X=OBS
    Só gerada quando projeto tem estacas disponíveis (processamento do zero).
    restricoes: ConfigRestricoes | None — quando corte_somente_bf, CL=0 na coluna Q.
    """
    if fatores_hom is None:
        fatores_hom = {}

    ws = wb.create_sheet("1_CV EIXOS", 0)  # primeira aba

    # Cabeçalho fixo linhas 1-8
    ws.merge_cells("A1:G1")
    ws["A1"] = nome_projeto or "MEMORIAL DE CÁLCULO DE VOLUMES"
    ws["A1"].font = Font(name="Arial", bold=True, size=10)

    # Linha 9 — cabeçalho de seções
    for col, txt in [(1,"ESTACA"), (2,"ÁREA SEÇÃO m²"), (7,"VOLUME PARCIAL GEOMÉTRICO m³"),
                     (12,"VOLUME PARCIAL HOMOGENEIZADO m³"), (17,"C.L"),
                     (18,"VOLUME HOMOGENEIZADO ACUMULADO m³"), (23,"O.M"), (24,"OBS.")]:
        c = ws.cell(9, col)
        c.value = txt
        c.font = Font(name="Arial", bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=COR_CABECALHO)
        c.font = Font(name="Arial", bold=True, size=9, color=FONTE_BRANCA)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borda_fina()

    # Linha 10 — subheader materiais
    mats = ["Corte 1ª", "Corte 2ª", "Corte 3ª", "Aterro CA", "Aterro CF"]
    for k, m in enumerate(mats):
        for base in [2, 7, 12, 18]:  # área, geo, hom, acumulado
            c = ws.cell(10, base + k)
            c.value = m
            c.font = Font(name="Arial", bold=True, size=8)
            c.fill = PatternFill("solid", fgColor=COR_SUBHEADER)
            c.alignment = Alignment(horizontal="center")
            c.border = borda_fina()

    # Larguras
    ws.column_dimensions["A"].width = 14
    for col in range(2, 25):
        ws.column_dimensions[get_column_letter(col)].width = 11

    # Mapeamento de materiais para categorias
    mat_c1 = mapeamento.corte1
    mat_c2 = mapeamento.corte2
    mat_c3 = mapeamento.corte3
    mat_ca = mapeamento.aterro_ca
    mat_cf = mapeamento.aterro_cf

    row = 11
    for ramo in projeto.ramos:
        if not ramo.estacas:
            continue

        # Título do ramo
        ws.merge_cells(f"A{row}:X{row}")
        ws[f"A{row}"] = ramo.nome
        ws[f"A{row}"].font = Font(name="Arial", bold=True, size=9)
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=COR_SUBHEADER)
        ws[f"A{row}"].border = borda_fina()
        row += 1

        # Verificar se ramo tem corte_somente_bf — nesse caso CL=0 (corte vai para BF)
        ramo_somente_bf = (restricoes is not None and
                           restricoes.tem_restricoes() and
                           restricoes.corte_somente_bf(ramo.nome))

        # Acumulados por categoria
        ac1 = ac2 = ac3 = aca = acf = 0.0
        om = 0.0

        for e in ramo.estacas:
            # Área seção (B-F)
            a_c1 = sum(e.areas.get(m, 0.0) for m in mat_c1)
            a_c2 = sum(e.areas.get(m, 0.0) for m in mat_c2)
            a_c3 = sum(e.areas.get(m, 0.0) for m in mat_c3)
            a_ca = sum(e.areas.get(m, 0.0) for m in mat_ca)
            a_cf = sum(e.areas.get(m, 0.0) for m in mat_cf)

            # Vol geométrico (G-K)
            g_c1 = sum(e.volumes.get(m, 0.0) for m in mat_c1)
            g_c2 = sum(e.volumes.get(m, 0.0) for m in mat_c2)
            g_c3 = sum(e.volumes.get(m, 0.0) for m in mat_c3)
            g_ca = sum(e.volumes.get(m, 0.0) for m in mat_ca)
            g_cf = sum(e.volumes.get(m, 0.0) for m in mat_cf)

            # Vol homogeneizado (L-P)
            h_c1 = sum(e.volumes_h.get(m, 0.0) for m in mat_c1)
            h_c2 = sum(e.volumes_h.get(m, 0.0) for m in mat_c2)
            h_c3 = sum(e.volumes_h.get(m, 0.0) for m in mat_c3)
            h_ca = sum(e.volumes_h.get(m, 0.0) for m in mat_ca)
            h_cf = sum(e.volumes_h.get(m, 0.0) for m in mat_cf)

            # CL por estaca (Q) = min(C1_hom, CA_hom)
            # Se ramo tem corte_somente_bf → CL=0 (corte vai para BF, não compensa aterro)
            if ramo_somente_bf:
                cl = 0.0
            else:
                cl = min(h_c1, h_ca) if h_c1 > 0 and h_ca > 0 else 0.0

            # Acumulados (R-V)
            ac1 += h_c1 - cl
            ac2 += h_c2
            ac3 += h_c3
            aca += h_ca - cl
            acf += h_cf

            # OM = Bruckner = acum_corte - acum_aterro
            om = round((ac1 + ac2 + ac3) - (aca + acf), 4)

            # Preencher linha
            fmt = '#,##0.00;-#,##0.00;"-"'
            ws.cell(row, 1).value = e.estaca

            for col, val in [(2,a_c1),(3,a_c2),(4,a_c3),(5,a_ca),(6,a_cf),
                             (7,g_c1),(8,g_c2),(9,g_c3),(10,g_ca),(11,g_cf),
                             (12,h_c1),(13,h_c2),(14,h_c3),(15,h_ca),(16,h_cf),
                             (17,cl),
                             (18,ac1),(19,ac2),(20,ac3),(21,aca),(22,acf),
                             (23,om)]:
                c = ws.cell(row, col)
                c.value = round(val, 4)  # sempre preenche com valor (0 vira traço pelo formato)
                c.number_format = fmt
                c.font = Font(name="Arial", size=8)
                c.border = borda_fina()

            ws.cell(row, 1).font = Font(name="Arial", size=8)
            ws.cell(row, 1).border = borda_fina()

            # Cor alternada por ramo
            row += 1

        row += 1  # linha em branco entre ramos

    return ws


# ---------------------------------------------------------------------------
# Função principal
# ---------------------------------------------------------------------------

def gerar_excel(resultado: ResultadoDistribuicao,
                resultados_deteccao: List[ResultadoDeteccao],
                caminho_saida: str,
                nome_projeto: str = "",
                fatores_hom: dict = None,
                mapeamento=None,
                projeto=None,
                restricoes=None):
    if fatores_hom is None:
        fatores_hom = {}
    wb = Workbook()
    wb.remove(wb.active)

    # Aba 1_CV EIXOS — só quando temos estacas disponíveis
    if projeto is not None and mapeamento is not None:
        tem_estacas = any(len(r.estacas) > 0 for r in projeto.ramos)
        if tem_estacas:
            print("  Gerando aba 1_CV EIXOS...")
            gerar_cv_eixos(wb, projeto, mapeamento, nome_projeto,
                           fatores_hom=fatores_hom, restricoes=restricoes)

    print("  Gerando aba 3_RESUMO_CA...")
    gerar_resumo_ca(wb, resultados_deteccao, nome_projeto,
                    fatores_hom=fatores_hom, restricoes=restricoes,
                    mapeamento=mapeamento)

    print("  Gerando aba 4_QDM...")
    ultima_linha = gerar_qdm(wb, resultado, resultados_deteccao)

    print(f"  Gerando aba 5_RESUMO_QDM (dados até linha {ultima_linha})...")
    gerar_resumo_qdm(wb, ultima_linha, fatores_hom=fatores_hom, mapeamento=mapeamento)

    wb.save(caminho_saida)
    print(f"  Excel salvo: {caminho_saida}")
    return caminho_saida