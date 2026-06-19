"""
leitor_civil.py
---------------
Lê arquivos Excel de terraplenagem em dois formatos:

Tipo 1 — Excel do Civil 3D (volumes já calculados)
  - Abas válidas: sem sufixo _PAR ou _XML, sem aba GRUPOS
  - Colunas: ESTACAO, Area X, Vol X, Vh X
  - Volumes homogeneizados já prontos

Tipo 2 — Excel só com áreas
  - Colunas: Estaca, Corte 1, Corte 2, Corte 3, Aterro PN, CF
  - Ferramenta calcula Vol_geo e Vol_hom usando Fh por material

Parâmetros comuns:
  - unidade: 'estaca' (÷20) ou 'km' (÷1000)
  - em_distancia: True = arquivo vem em metros, converter para estaca/km
  - dist_estaca: distância entre estacas em metros (padrão 20m) — Tipo 2
  - fatores_hom: dict {material: Fh} — Tipo 2
"""

import os
from dataclasses import dataclass, field
from typing import Dict, List, Optional
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Estruturas de dados
# ---------------------------------------------------------------------------

@dataclass
class DadosEstaca:
    estaca:    str              # label formatado ex: '117+10,000'
    estaca_m:  float            # posição em metros (sempre)
    areas:     Dict[str, float] = field(default_factory=dict)
    volumes:   Dict[str, float] = field(default_factory=dict)   # geométrico
    volumes_h: Dict[str, float] = field(default_factory=dict)   # homogeneizado
    cl:        float = 0.0


@dataclass
class DadosRamo:
    nome:       str
    nome_aba:   str
    arquivo:    str
    tipo:       int             # 1 = Civil 3D, 2 = só áreas
    materiais:  List[str] = field(default_factory=list)
    estacas:    List[DadosEstaca] = field(default_factory=list)


@dataclass
class DadosProjeto:
    ramos: List[DadosRamo] = field(default_factory=list)


@dataclass
class ConfigLeitura:
    """Parâmetros de leitura informados pelo usuário."""
    unidade:        str   = "estaca"  # 'estaca' ou 'km'
    em_distancia:   bool  = False     # True = arquivo em metros
    dist_estaca:    float = 20.0      # metros entre estacas (Tipo 2)
    fatores_hom:    Dict[str, float] = field(default_factory=dict)  # Tipo 2
    dist_max_bloco: float = 20.0      # distância máxima entre estacas consecutivas
                                      # acima disso o volume entre elas é zerado
                                      # (blocos separados sem material contínuo)


# ---------------------------------------------------------------------------
# Conversão de posição
# ---------------------------------------------------------------------------

def posicao_para_metros(valor: str, config: ConfigLeitura) -> float:
    """
    Converte o valor de posição do arquivo para metros internos.

    Se em_distancia=True: valor já está em metros → usa direto
    Se em_distancia=False:
      unidade='estaca': 'NNN+RR,RRR' → (NNN × 20) + RR,RRR
      unidade='km':     'K+MMM,MMM'  → (K × 1000) + MMM,MMM
    """
    try:
        s = str(valor).strip().replace(",", ".")
        if config.em_distancia:
            return float(s)

        if "+" in s:
            partes = s.split("+")
            num    = float(partes[0])
            resto  = float(partes[1])
            if config.unidade == "km":
                return num * 1000.0 + resto
            else:
                return num * config.dist_estaca + resto
        else:
            return float(s)
    except Exception:
        return 0.0


def metros_para_label(metros: float, config: ConfigLeitura) -> str:
    """Formata metros para label de estaca/km."""
    if config.unidade == "km":
        km    = int(metros // 1000)
        resto = metros % 1000
        return f"{km}+{resto:07.3f}".replace(".", ",")
    else:
        num   = int(metros // config.dist_estaca)
        resto = metros % config.dist_estaca
        return f"{num}+{resto:06.3f}".replace(".", ",")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_float(v) -> float:
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0


def is_aba_valida(nome: str) -> bool:
    u = nome.upper()
    if u == "GRUPOS": return False
    if u.endswith("_PAR"): return False
    if u.endswith("_XML"): return False
    return True


def extrair_nome_ramo(nome_aba: str, nome_arquivo: str) -> str:
    """Remove prefixo 'SEÇÕES - ' se existir, senão usa o próprio nome da aba."""
    prefixo = "SEÇÕES - "
    if nome_aba.upper().startswith(prefixo.upper()):
        return nome_aba[len(prefixo):]
    # Usa o nome da aba diretamente como nome do ramo
    return nome_aba


# ---------------------------------------------------------------------------
# Leitura Tipo 1 — Excel Civil 3D
# ---------------------------------------------------------------------------

def ler_aba_tipo1(ws, nome_arquivo: str,
                  config: ConfigLeitura) -> Optional[DadosRamo]:
    """Lê uma aba do Excel do Civil 3D (volumes já calculados)."""
    nome_aba  = ws.title
    nome_ramo = extrair_nome_ramo(nome_aba, nome_arquivo)
    ramo = DadosRamo(nome=nome_ramo, nome_aba=nome_aba,
                     arquivo=nome_arquivo, tipo=1)

    linhas = list(ws.iter_rows(values_only=True))
    if len(linhas) < 2:
        return None

    # Encontrar cabeçalho (linha com "ESTACAO")
    idx_cab = None
    for i, linha in enumerate(linhas):
        if linha and any(str(c).upper().strip() == "ESTACAO"
                         for c in linha if c is not None):
            idx_cab = i
            break

    if idx_cab is None:
        print(f"  AVISO: cabeçalho não encontrado em '{nome_aba}'")
        return None

    cab = linhas[idx_cab]
    colunas: Dict[str, int] = {}
    for i, c in enumerate(cab):
        if c is not None:
            colunas[str(c).strip()] = i

    # Identificar materiais (colunas Area X, Vol X, Vh X)
    materiais_area = [c.replace("Area ", "") for c in colunas if c.startswith("Area ")]
    materiais_vol  = [c.replace("Vol ", "")  for c in colunas if c.startswith("Vol ")]
    materiais_vh   = [c.replace("Vh ", "")   for c in colunas if c.startswith("Vh ")]
    materiais = [m for m in materiais_area if m in materiais_vol and m in materiais_vh]
    ramo.materiais = materiais

    idx_cl  = colunas.get("C.L.", None)
    idx_est = colunas.get("ESTACAO", 0)

    for linha in linhas[idx_cab + 1:]:
        if not linha or linha[idx_est] is None:
            continue
        est_raw = str(linha[idx_est]).strip()
        if not est_raw:
            continue

        est_m   = posicao_para_metros(est_raw, config)
        est_lbl = metros_para_label(est_m, config)

        dados = DadosEstaca(estaca=est_lbl, estaca_m=est_m)
        dados.cl = safe_float(linha[idx_cl]) if idx_cl else 0.0

        for mat in materiais:
            dados.areas[mat]     = safe_float(linha[colunas.get(f"Area {mat}", -1)]) \
                                   if f"Area {mat}" in colunas else 0.0
            dados.volumes[mat]   = safe_float(linha[colunas.get(f"Vol {mat}", -1)]) \
                                   if f"Vol {mat}" in colunas else 0.0
            dados.volumes_h[mat] = safe_float(linha[colunas.get(f"Vh {mat}", -1)]) \
                                   if f"Vh {mat}" in colunas else 0.0

        # Verificar salto de bloco — se distância > dist_max_bloco zera volumes desta estaca
        if ramo.estacas:
            dist_ant = est_m - ramo.estacas[-1].estaca_m
            if dist_ant > config.dist_max_bloco:
                for mat in materiais:
                    dados.volumes[mat]   = 0.0
                    dados.volumes_h[mat] = 0.0

        ramo.estacas.append(dados)

    return ramo


# ---------------------------------------------------------------------------
# Leitura Tipo 2 — Excel só com áreas
# ---------------------------------------------------------------------------

def ler_aba_tipo2(ws, nome_arquivo: str,
                  config: ConfigLeitura) -> Optional[DadosRamo]:
    """
    Lê Excel com apenas áreas de seção por material.
    Calcula volumes geométricos e homogeneizados.
    Formato esperado: Estaca | Corte 1 | Corte 2 | Corte 3 | Aterro PN | CF
    """
    nome_aba  = ws.title
    nome_ramo = extrair_nome_ramo(nome_aba, nome_arquivo)
    ramo = DadosRamo(nome=nome_ramo, nome_aba=nome_aba,
                     arquivo=nome_arquivo, tipo=2)

    linhas = list(ws.iter_rows(values_only=True))
    if len(linhas) < 2:
        return None

    # Encontrar cabeçalho — linha que contém nome de estaca OU primeira linha com dados numéricos
    idx_cab = None
    for i, linha in enumerate(linhas):
        if not linha:
            continue
        # Verificar se alguma célula tem nome de estaca
        if any(c is not None and
               str(c).upper().strip() in ("ESTACA", "ESTACAO", "STATION")
               for c in linha):
            idx_cab = i
            break
        # Verificar se linha tem células não-nulas com texto (possível cabeçalho)
        # e a próxima linha tem valores numéricos ou de estaca
        celulas_texto = [c for c in linha if c is not None and isinstance(c, str) and c.strip()]
        if celulas_texto and i + 1 < len(linhas):
            proxima = linhas[i + 1]
            if proxima and any(c is not None for c in proxima):
                idx_cab = i
                break

    if idx_cab is None:
        print(f"  AVISO: cabeçalho não encontrado em '{nome_aba}'")
        return None

    cab = linhas[idx_cab]

    # Mapear colunas de material — usa os nomes originais das colunas como chave
    # Identifica a coluna de estaca e todas as demais são materiais
    mapa_col: Dict[str, int] = {}  # nome_original → índice
    col_estaca = None

    for i, c in enumerate(cab):
        if c is None:
            # Coluna sem cabeçalho — verifica se é a primeira coluna (provável estaca)
            if i == 0 and col_estaca is None:
                col_estaca = 0
                mapa_col["ESTACA"] = 0
            continue
        c_str = str(c).strip()
        c_up  = c_str.upper()
        if c_up in ("ESTACA", "ESTACAO", "STATION") or (i == 0 and col_estaca is None):
            col_estaca = i
            mapa_col["ESTACA"] = i
        else:
            # Toda coluna não-estaca é um material — usa nome original
            if c_str:
                mapa_col[c_str] = i

    # Se ainda não achou coluna estaca, usar coluna 0
    if col_estaca is None:
        col_estaca = 0
        mapa_col["ESTACA"] = 0

    # Materiais = todas as colunas exceto ESTACA
    materiais = [k for k in mapa_col if k != "ESTACA"]
    ramo.materiais = materiais

    # Aplicar FH dos fatores_hom (usuário informa na Janela 3)
    # Se não informado, usa 1.0 (sem homogeneização — volumes geométricos)

    # Ler estacas
    estacas_raw: List[DadosEstaca] = []
    for linha in linhas[idx_cab + 1:]:
        if not linha or linha[mapa_col.get("ESTACA", 0)] is None:
            continue
        est_raw = str(linha[mapa_col["ESTACA"]]).strip()
        if not est_raw:
            continue

        est_m   = posicao_para_metros(est_raw, config)
        est_lbl = metros_para_label(est_m, config)
        dados   = DadosEstaca(estaca=est_lbl, estaca_m=est_m)

        for mat in materiais:
            dados.areas[mat] = safe_float(linha[mapa_col[mat]])

        estacas_raw.append(dados)

    # Calcular volumes incrementais (seções médias)
    for i, e in enumerate(estacas_raw):
        if i == 0:
            for mat in materiais:
                e.volumes[mat]   = 0.0
                e.volumes_h[mat] = 0.0
        else:
            e_ant = estacas_raw[i - 1]
            dist  = e.estaca_m - e_ant.estaca_m
            if dist <= 0:
                dist = config.dist_estaca

            # Se a distância entre estacas consecutivas excede dist_max_bloco
            # são blocos separados — zera o volume entre elas
            if dist > config.dist_max_bloco:
                for mat in materiais:
                    e.volumes[mat]   = 0.0
                    e.volumes_h[mat] = 0.0
            else:
                for mat in materiais:
                    area_med = (e_ant.areas.get(mat, 0.0) + e.areas.get(mat, 0.0)) / 2.0
                    vol_geo  = round(area_med * dist, 4)
                    fh       = config.fatores_hom.get(mat, 1.0)
                    # FH divide o volume geométrico para obter o volume homogeneizado
                    vol_hom  = round(vol_geo / fh, 4) if fh > 0 else vol_geo
                    e.volumes[mat]   = vol_geo
                    e.volumes_h[mat] = vol_hom

        ramo.estacas.append(e)

    return ramo


# ---------------------------------------------------------------------------
# Funções principais de leitura
# ---------------------------------------------------------------------------

def ler_excel_civil(caminho: str,
                    config: Optional[ConfigLeitura] = None) -> DadosProjeto:
    """
    Lê um Excel do Civil 3D (Tipo 1).
    Pode ser chamado com múltiplos arquivos via ler_multiplos_arquivos.
    """
    if config is None:
        config = ConfigLeitura()

    projeto = DadosProjeto()

    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    print(f"\nLendo (Tipo 1): {os.path.basename(caminho)}")
    wb = load_workbook(caminho, read_only=True, data_only=True)

    abas_validas = [n for n in wb.sheetnames if is_aba_valida(n)]
    print(f"  Abas válidas: {abas_validas}")

    for nome_aba in abas_validas:
        ws   = wb[nome_aba]
        ramo = ler_aba_tipo1(ws, caminho, config)
        if ramo and ramo.estacas:
            print(f"  → {ramo.nome}: {len(ramo.estacas)} estacas, "
                  f"materiais: {ramo.materiais}")
            projeto.ramos.append(ramo)

    wb.close()
    return projeto


def ler_excel_areas(caminho: str,
                    config: Optional[ConfigLeitura] = None) -> DadosProjeto:
    """
    Lê um Excel só com áreas (Tipo 2) e calcula volumes.
    """
    if config is None:
        config = ConfigLeitura()

    projeto = DadosProjeto()

    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    print(f"\nLendo (Tipo 2): {os.path.basename(caminho)}")
    wb = load_workbook(caminho, read_only=True, data_only=True)

    for nome_aba in wb.sheetnames:
        ws   = wb[nome_aba]
        ramo = ler_aba_tipo2(ws, caminho, config)
        if ramo and ramo.estacas:
            print(f"  → {ramo.nome}: {len(ramo.estacas)} estacas, "
                  f"materiais: {ramo.materiais}")
            projeto.ramos.append(ramo)

    wb.close()
    return projeto


def ler_multiplos_arquivos(arquivos: List[Dict],
                            config: Optional[ConfigLeitura] = None) -> DadosProjeto:
    """
    Lê múltiplos arquivos Excel e consolida em um único DadosProjeto.

    arquivos: lista de dicts com:
      - 'caminho': str
      - 'tipo': 1 (Civil 3D) ou 2 (só áreas)

    Exemplo:
      arquivos = [
        {'caminho': 'RAMO100.xlsx', 'tipo': 1},
        {'caminho': 'RAMO300.xlsx', 'tipo': 1},
        {'caminho': 'SEGMENTO_A.xlsx', 'tipo': 2},
      ]
    """
    if config is None:
        config = ConfigLeitura()

    projeto_total = DadosProjeto()

    for arq in arquivos:
        caminho = arq.get("caminho", "")
        tipo    = arq.get("tipo", 1)

        if tipo == 1:
            p = ler_excel_civil(caminho, config)
        else:
            p = ler_excel_areas(caminho, config)

        projeto_total.ramos.extend(p.ramos)

    print(f"\nTotal de ramos carregados: {len(projeto_total.ramos)}")
    for r in projeto_total.ramos:
        print(f"  - {r.nome} ({len(r.estacas)} estacas) "
              f"[Tipo {r.tipo}] [{r.arquivo}]")

    return projeto_total


def imprimir_resumo(projeto: DadosProjeto):
    """Imprime resumo dos dados lidos para conferência."""
    print(f"\n{'='*60}")
    print("RESUMO DOS DADOS LIDOS")
    print(f"{'='*60}")
    print(f"Total de ramos: {len(projeto.ramos)}")

    for ramo in projeto.ramos:
        print(f"\n{'─'*40}")
        print(f"RAMO: {ramo.nome} [Tipo {ramo.tipo}]")
        print(f"  Arquivo:   {os.path.basename(ramo.arquivo)}")
        print(f"  Materiais: {', '.join(ramo.materiais)}")
        print(f"  Estacas:   {len(ramo.estacas)}")

        if not ramo.estacas:
            continue

        print(f"  De: {ramo.estacas[0].estaca} → Até: {ramo.estacas[-1].estaca}")
        print(f"\n  {'Material':<15} {'Vol Geom (m³)':>15} {'Vol Hom (m³)':>15}")
        print(f"  {'─'*45}")

        for mat in ramo.materiais:
            vg = sum(e.volumes.get(mat, 0) for e in ramo.estacas)
            vh = sum(e.volumes_h.get(mat, 0) for e in ramo.estacas)
            if vg > 0 or vh > 0:
                print(f"  {mat:<15} {vg:>15,.2f} {vh:>15,.2f}")

        print(f"\n  Primeiras 3 estacas:")
        for e in ramo.estacas[:3]:
            vols = {m: v for m, v in e.volumes_h.items() if v > 0}
            print(f"    {e.estaca} | Vh: {vols}")

        print(f"\n  Últimas 3 estacas:")
        for e in ramo.estacas[-3:]:
            vols = {m: v for m, v in e.volumes_h.items() if v > 0}
            print(f"    {e.estaca} | Vh: {vols}")


# ---------------------------------------------------------------------------
# Execução direta para teste
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    config = ConfigLeitura(
        unidade      = "estaca",
        em_distancia = False,
        dist_estaca  = 20.0,
        fatores_hom  = {
            "Corte1": 1.00,
            "Corte2": 1.15,
            "Corte3": 0.90,
            "Aterro": 1.25,
            "CF":     1.25,
        }
    )

    arquivos = [
        {"caminho": r"E:\Dropbox\VisualStudio\PROGRAMA PARA DISTRIBUIÇÃO\DISTRIBUICAO\DISTRIBUICAO\SEGB_INT 03_Civil.xlsx",
         "tipo": 1},
    ]

    projeto = ler_multiplos_arquivos(arquivos, config)
    imprimir_resumo(projeto)